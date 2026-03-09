# -*- coding: utf-8 -*-
# 指定文件编码为 UTF-8，以便在源代码中使用中文
"""
================================================================================
                    openpyxl 入门教程
================================================================================

openpyxl 是 Python 中读写 Excel 2007+ (.xlsx) 格式的常用库。
相比 pandas，它更擅长：单元格级操作、样式、公式、合并单元格等。

安装：  pip install openpyxl

本教程按顺序演示：创建、读取、写入、样式、常用 API。
================================================================================
"""

# =============================================================================
# 1. 创建新的 Excel 工作簿
# =============================================================================

from openpyxl import Workbook
# 从 openpyxl 包中导入 Workbook 类，用于创建新的 Excel 工作簿

# # 创建一个空白工作簿（相当于新建一个 Excel 文件）
wb = Workbook()
# 实例化 Workbook 对象，相当于在内存中新建了一个空的 .xlsx 文件

# # 获取默认的工作表（Excel 里叫 Sheet）
ws = wb.active
# active 属性返回当前活动的工作表，新建工作簿时默认有一个名为 "Sheet" 的表

ws.title = "员工表"  # 可以改工作表名称
# 设置工作表的名称，会显示在 Excel 底部的工作表标签上

# 写入单个单元格：ws['A1'] = 值
ws["A1"] = "姓名"
# 通过单元格地址（如 A1）直接赋值，将字符串 "姓名" 写入 A1 单元格

ws["B1"] = "部门"
# 将 "部门" 写入 B1 单元格

ws["C1"] = "工资"
# 将 "工资" 写入 C1 单元格，完成表头第一行的写入

# # 写入多行数据
ws["A2"] = "安欣"
# 第 2 行 A 列：员工姓名

ws["B2"] = "刑警队"
# 第 2 行 B 列：所属部门

ws["C2"] = 12000
# 第 2 行 C 列：工资数值（整数类型）

ws["A3"] = "高启强"
# 第 3 行 A 列

ws["B3"] = "建工集团"
# 第 3 行 B 列

ws["C3"] = 30000
# 第 3 行 C 列

# # 也可以按行列号写入：ws.cell(row=行, column=列, value=值)
ws.cell(row=4, column=1, value="李响")
# cell() 方法通过行列号访问单元格，row=4 表示第 4 行，column=1 表示第 1 列（A 列）

ws.cell(row=4, column=2, value="刑警队")
# 第 4 行第 2 列（B 列）写入部门

ws.cell(row=4, column=3, value=13000)
# 第 4 行第 3 列（C 列）写入工资

# # 保存到文件
wb.save("openpyxl_示例_创建.xlsx")
# 将工作簿保存为指定路径的 .xlsx 文件，会覆盖同名文件

print("已保存: openpyxl_示例_创建.xlsx")
# 在控制台输出保存成功的提示信息

# =============================================================================
# 2. 批量写入数据（append / 切片）
# =============================================================================

from openpyxl import Workbook
# 导入 Workbook 类，用于创建新的 Excel 工作簿

wb = Workbook()
# 创建一个空白工作簿对象，相当于新建一个 Excel 文件

ws = wb.active
# 获取当前活动的工作表，即默认的第一个 Sheet

ws.title = "工资明细"
# 将工作表名称改为 "工资明细"，便于识别

# 表头
headers = ["姓名", "部门", "基础工资", "绩效奖金", "扣款", "实发工资", "额外奖金"]
# 定义表头列表，每个元素对应一列的列名，顺序从左到右依次为 A、B、C... 列

ws.append(headers)
# append() 方法将列表作为一行追加到工作表末尾，这里把表头写入第一行

# 用 append() 逐行追加（每行一个列表）
rows = [
    # 定义数据行列表，每个内层列表代表一行数据
    ["安欣", "刑警队", 12000, 5000, 800, 16200, 200],
    # 第 1 条记录：姓名、部门、基础工资、绩效奖金、扣款、实发工资、额外奖金

    ["高启强", "建工集团", 30000, 8000, 3000, 35000, 100],
    # 第 2 条记录

    ["李响", "刑警队", 13000, 4000, 500, 16500, 50],
    # 第 3 条记录
]
for row in rows:
    # 遍历 rows 中的每一行
    ws.append(row)
    # 将当前行（列表）追加到工作表，自动从第 2 行开始（表头已在第 1 行）

wb.save("openpyxl_示例_批量写入.xlsx")
# 将工作簿保存到当前目录下的指定文件名

print("已保存: openpyxl_示例_批量写入.xlsx")
# 打印保存成功的提示

# =============================================================================
# 3. 读取已有的 Excel 文件
# =============================================================================

from openpyxl import load_workbook
# 导入 load_workbook 函数，用于加载已存在的 Excel 文件

wb = load_workbook("openpyxl_示例_批量写入.xlsx")
# 加载已存在的 Excel 文件（可读写），打开指定路径的 .xlsx 文件，返回 Workbook 对象

ws = wb.active  # 第一个工作表
# 获取当前活动工作表（通常是第一个）
# ws = wb["工资明细"]  # 也可以按名称获取

print("\n--- 读取单元格 ---")
# 打印分隔线

print("A1:", ws["A1"].value)
# 读取 A1 单元格的值，.value 属性获取单元格中的实际数据

print("B2:", ws["B2"].value)
# 读取 B2 单元格的值

print("(2,3):", ws.cell(row=2, column=3).value)
# 按行列号读取，通过 cell(row=行, column=列) 获取单元格，再取 .value

print("\n--- 姓名列 ---")
# 遍历某一列

for row in range(2, ws.max_row + 1):
    # 从第 2 行遍历到最大行，ws.max_row 是工作表中有数据的最大行号
    name = ws.cell(row=row, column=1).value
    # 获取当前行第 1 列（A 列）的值，即姓名
    print(name)

print("\n--- 所有数据行 ---")
# 遍历所有数据（跳过表头）

for row in ws.iter_rows(min_row=2, values_only=True):
    # iter_rows() 按行迭代，min_row=2 从第 2 行开始跳过表头，values_only=True 只返回值
    print(row)

# =============================================================================
# 4. 修改单元格 + 样式
# =============================================================================

from openpyxl import load_workbook
# 导入加载工作簿的函数

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# 从 openpyxl.styles 导入样式相关类：Font 字体、Alignment 对齐、Border 边框、Side 边框边、PatternFill 填充

wb = load_workbook("openpyxl_示例_批量写入.xlsx")
# 加载之前保存的批量写入示例文件

ws = wb.active
# 获取活动工作表

ws["A1"] = "员工姓名"  # 改表头
# 修改单元格内容，将 A1 的 "姓名" 改为 "员工姓名"

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
# 创建填充对象：蓝色背景，4472C4 为十六进制颜色码，fill_type="solid" 表示纯色填充

header_alignment = Alignment(horizontal="center", vertical="center")
# 创建对齐对象：水平居中、垂直居中

for col in range(1, 8):  # A 到 G 列（7 列，与表头对应）
    # 遍历第 1 到 7 列
    cell = ws.cell(row=1, column=col)
    # 获取表头行（第 1 行）的当前列单元格
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    # 设置字体：加粗、字号 12、白色，用于在深色背景上显示
    cell.fill = header_fill
    # 设置单元格背景填充为蓝色
    cell.alignment = header_alignment
    # 设置单元格内容居中对齐

ws.column_dimensions["A"].width = 12
# 设置 A 列宽度为 12 个字符单位

ws.column_dimensions["B"].width = 12
# 设置 B 列宽度

ws.column_dimensions["C"].width = 12
# 设置 C 列宽度

thin_border = Border(
    # 创建边框对象，指定四条边的样式
    left=Side(style="thin"),
    # 左边框：细线样式
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
    # 遍历第 1 行到最大行、第 1 到 7 列的所有单元格区域
    for cell in row:
        # 对区域中的每一行，遍历该行的每个单元格
        cell.border = thin_border
        # 为每个单元格应用细边框

wb.save("openpyxl_示例_带样式.xlsx")
# 保存到新文件，避免覆盖原文件

print("\n已保存: openpyxl_示例_带样式.xlsx")
# 打印保存提示

# =============================================================================
# 5. 多工作表操作
# =============================================================================

from openpyxl import Workbook
# 导入 Workbook

wb = Workbook()
# 创建新工作簿

wb.remove(wb.active)
# 删除默认的 Sheet，remove() 删除指定工作表

ws1 = wb.create_sheet("工资明细", 0)
# create_sheet(名称, 位置)：创建名为 "工资明细" 的工作表，0 表示插在第一位

ws2 = wb.create_sheet("部门统计", 1)
# 创建 "部门统计" 表，插入在第二个位置

ws3 = wb.create_sheet("说明", 2)
# 创建 "说明" 表，插入在第三个位置

ws1["A1"] = "工资明细表"
# 在第一个工作表的 A1 写入标题

ws2["A1"] = "部门统计表"
# 在第二个工作表的 A1 写入标题

ws3["A1"] = "本文件包含多个工作表"
# 在第三个工作表的 A1 写入说明文字

wb.save("openpyxl_示例_多工作表.xlsx")
# 保存工作簿，一个 .xlsx 文件可包含多个工作表

print("已保存: openpyxl_示例_多工作表.xlsx")
# 打印提示

# =============================================================================
# 6. 合并单元格
# =============================================================================

from openpyxl import Workbook
# 导入 Workbook

from openpyxl.styles import Alignment
# 导入 Alignment 用于设置合并后单元格的对齐方式

wb = Workbook()
# 创建新工作簿

ws = wb.active
# 获取活动工作表

ws["A1"] = "工资条汇总表"
# 在 A1 写入标题文字

ws.merge_cells("A1:F1")  # 合并 A1 到 F1
# merge_cells() 合并指定区域的单元格，将第一行 A 到 F 列合并为一个单元格

ws["A1"].alignment = Alignment(horizontal="center")
# 合并后只有左上角单元格保留内容，设置其水平居中

ws["A2"] = "姓名"
# 在 A2 写入 "姓名"

ws["B2"] = "部门"
# 在 B2 写入 "部门"

ws.merge_cells("A2:B2")  # 合并 A2:B2（示例用）
# 合并 A2 和 B2，演示合并相邻单元格的用法

wb.save("openpyxl_示例_合并单元格.xlsx")
# 保存文件

print("已保存: openpyxl_示例_合并单元格.xlsx")
# 打印提示

# =============================================================================
# 7. 常用 API 速查
# =============================================================================
"""
┌─────────────────────────────────────────────────────────────────┐
│  openpyxl 常用 API 速查                                          │
├─────────────────────────────────────────────────────────────────┤
│  from openpyxl import Workbook, load_workbook                    │
│  导入：Workbook 用于创建，load_workbook 用于加载已有文件          │
│                                                                  │
│  创建/加载：                                                      │
│    wb = Workbook()           # 新建空白工作簿                     │
│    wb = load_workbook(path)  # 加载已有 .xlsx 文件                │
│                                                                  │
│  工作表：                                                         │
│    ws = wb.active            # 获取当前活动表                     │
│    ws = wb["Sheet1"]         # 按工作表名称获取                   │
│    wb.create_sheet("新表")   # 新建工作表                         │
│    ws.title = "新名称"       # 修改工作表名称                     │
│                                                                  │
│  读写单元格：                                                     │
│    ws["A1"] = 100            # 按地址写入                         │
│    val = ws["A1"].value      # 按地址读取，.value 取实际值        │
│    ws.cell(row=1, col=1, value=100)  # 按行列号读写               │
│                                                                  │
│  批量：                                                           │
│    ws.append([a, b, c])      # 在末尾追加一行                     │
│    ws["A1":"C3"]             # 区域切片，获取 A1 到 C3 范围       │
│                                                                  │
│  遍历：                                                           │
│    ws.iter_rows(min_row=1, max_row=10, values_only=True)         │
│    # 按行迭代，values_only=True 只返回值不返回 Cell 对象          │
│    ws.max_row, ws.max_column # 获取有数据的最大行号、列号         │
│                                                                  │
│  样式（from openpyxl.styles）：                                   │
│    Font, Alignment, Border, PatternFill  # 字体、对齐、边框、填充 │
│                                                                  │
│  保存：                                                           │
│    wb.save("xxx.xlsx")       # 保存到文件，会覆盖同名文件         │
└─────────────────────────────────────────────────────────────────┘
"""
